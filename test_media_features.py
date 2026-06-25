import asyncio
import importlib
import sys
import tempfile
import types
import unittest
import warnings
from pathlib import Path

from openpyxl import load_workbook


class _FakeCuda:
    @staticmethod
    def is_available():
        return False


sys.modules.setdefault(
    "torch",
    types.SimpleNamespace(__version__="test", cuda=_FakeCuda()),
)
sys.modules.setdefault(
    "whisper",
    types.SimpleNamespace(load_model=lambda *args, **kwargs: object()),
)

main = importlib.import_module("main")
db = importlib.import_module("database")
warnings.simplefilter("ignore", ResourceWarning)


class MediaFeatureTests(unittest.TestCase):
    def test_parse_srt_handles_multiline_cues(self):
        cues = main.parse_srt(
            "1\n"
            "00:00:01,000 --> 00:00:03,500\n"
            "첫 줄\n"
            "둘째 줄\n\n"
            "2\n"
            "00:00:04,000 --> 00:00:05,000\n"
            "다음 자막\n"
        )

        self.assertEqual(len(cues), 2)
        self.assertEqual(cues[0]["index"], 1)
        self.assertEqual(cues[0]["start"], 1.0)
        self.assertEqual(cues[0]["end"], 3.5)
        self.assertEqual(cues[0]["text"], "첫 줄\n둘째 줄")

    def test_parse_srt_skips_invalid_timecode(self):
        cues = main.parse_srt(
            "1\n"
            "bad timecode\n"
            "무시됨\n\n"
            "2\n"
            "00:00:02,000 --> 00:00:04,000\n"
            "유효함\n"
        )

        self.assertEqual(len(cues), 1)
        self.assertEqual(cues[0]["index"], 2)

    def test_build_srt_preserves_index_and_timecodes(self):
        cues = main.parse_srt(
            "7\n"
            "00:01:02,003 --> 00:01:04,900\n"
            "원문\n"
        )

        output = main.build_srt(cues, ["Translated text"])

        self.assertIn("7", output)
        self.assertIn("00:01:02,003 --> 00:01:04,900", output)
        self.assertIn("Translated text", output)

    def test_build_ass_subtitles_applies_style(self):
        ass_text, normalized = main.build_ass_subtitles(
            "1\n00:00:01,000 --> 00:00:03,500\nHello\nWorld",
            main.SubtitleStyleRequest(
                font_family="Malgun Gothic",
                font_size=54,
                position="top",
                text_color="#ffee00",
                background_enabled=True,
                background_color="#112233",
                background_opacity=70,
                outline_width=4,
                shadow=2,
                margin_v=32,
            ),
        )

        self.assertEqual(normalized["position"], "top")
        self.assertIn("Style: Default,Malgun Gothic,54", ass_text)
        self.assertIn("&H0000EEFF", ass_text)
        self.assertIn("&H4D332211", ass_text)
        self.assertIn(",3,4,2,8,80,80,32,1", ass_text)
        self.assertIn("Dialogue: 0,0:00:01.00,0:00:03.50", ass_text)
        self.assertIn(r"Hello\NWorld", ass_text)

    def test_translate_srt_to_english_preserves_timing(self):
        class _FakeResponse:
            text = '[{"index": 1, "text": "Hello"}, {"index": 2, "text": "World"}]'

        class _FakeModels:
            @staticmethod
            def generate_content(model, contents):
                return _FakeResponse()

        class _FakeGeminiClient:
            models = _FakeModels()

        original_client = main.gemini_text_client
        original_key = main.gemini_api_key
        main.gemini_text_client = _FakeGeminiClient()
        main.gemini_api_key = "test-key"
        try:
            output = asyncio.run(
                main.translate_srt_to_english(
                    "1\n"
                    "00:00:00,000 --> 00:00:01,500\n"
                    "안녕\n\n"
                    "2\n"
                    "00:00:02,000 --> 00:00:03,000\n"
                    "세계\n"
                )
            )
        finally:
            main.gemini_text_client = original_client
            main.gemini_api_key = original_key

        self.assertIn("00:00:00,000 --> 00:00:01,500", output)
        self.assertIn("00:00:02,000 --> 00:00:03,000", output)
        self.assertIn("Hello", output)
        self.assertIn("World", output)

    def test_srt_tts_timeline_matches_last_cue_end(self):
        calls = []

        async def fake_generate_tts_wav(text, language, wav_path, voice_name=None, style_prompt=None, tts_provider=None):
            calls.append(text)
            main.write_wave_file(Path(wav_path), b"\0\0" * 2400)

        original_generate_tts_wav = main.generate_tts_wav
        main.generate_tts_wav = fake_generate_tts_wav
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                output_path = Path(temp_dir) / "timed.mp3"
                metadata = asyncio.run(
                    main.synthesize_audio_from_srt(
                        "1\n"
                        "00:00:00,000 --> 00:00:01,000\n"
                        "Hello\n\n"
                        "2\n"
                        "00:00:01,500 --> 00:00:03,000\n"
                        "World\n",
                        "en",
                        output_path,
                    )
                )
                audio = main.AudioSegment.from_file(output_path)
        finally:
            main.generate_tts_wav = original_generate_tts_wav

        self.assertEqual(metadata["cue_count"], 2)
        self.assertEqual(metadata["tts_request_count"], 2)
        self.assertEqual(metadata["tts_sync_mode"], "cue")
        self.assertEqual(len(calls), 2)
        self.assertLess(abs(metadata["duration_ms"] - 3000), 80)
        self.assertLess(abs(len(audio) - 3000), 150)

    def test_file_detail_includes_artifact_summary(self):
        original_db_path = db.DB_PATH
        original_json_path = db.JSON_DB_PATH
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                db.DB_PATH = Path(temp_dir) / "files.sqlite"
                db.JSON_DB_PATH = Path(temp_dir) / "missing.json"
                db.init_db()
                file_record = db.create_file_record(
                    "sample.mp4",
                    "video",
                    "text",
                    srt_text="1\n00:00:00,000 --> 00:00:01,000\n안녕",
                    media_path="media/sample.mp4",
                    english_srt_text="1\n00:00:00,000 --> 00:00:01,000\nHello",
                )
                db.create_artifact(
                    file_record["id"],
                    "audio",
                    "en",
                    "outputs/sample_en.mp3",
                    "sample_en.mp3",
                    {"duration_ms": 1000},
                )

                detail = db.get_file_by_id(file_record["id"])
        finally:
            db.DB_PATH = original_db_path
            db.JSON_DB_PATH = original_json_path

        self.assertTrue(detail["artifact_summary"]["audio_en"])
        self.assertFalse(detail["artifact_summary"]["video_en"])
        self.assertEqual(detail["artifacts"][0]["filename"], "sample_en.mp3")

    def test_dub_video_reuses_existing_audio_artifact(self):
        original_db_path = db.DB_PATH
        original_json_path = db.JSON_DB_PATH
        original_mux = main.mux_video_with_audio
        original_black_video = main.create_black_video
        original_synthesize = main.synthesize_audio_from_srt
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                db.DB_PATH = Path(temp_dir) / "files.sqlite"
                db.JSON_DB_PATH = Path(temp_dir) / "missing.json"
                db.init_db()
                audio_path = Path(temp_dir) / "existing.mp3"
                audio_path.write_bytes(b"fake mp3")
                file_record = db.create_file_record(
                    "sample.srt",
                    "srt_project",
                    "text",
                    srt_text="1\n00:00:00,000 --> 00:00:01,000\nHello",
                )
                db.create_artifact(
                    file_record["id"],
                    "audio",
                    "ko",
                    str(audio_path),
                    "existing.mp3",
                    {"duration_ms": 1000, "voice_name": "Kore", "srt_source": "original", "tts_sync_mode": "cue"},
                )

                def fake_black_video(duration_ms, output_path):
                    Path(output_path).write_bytes(b"fake video")

                def fake_mux(video_path, input_audio_path, output_path):
                    self.assertEqual(Path(input_audio_path), audio_path)
                    Path(output_path).write_bytes(b"muxed")

                async def fail_synthesize(*args, **kwargs):
                    raise AssertionError("TTS should not be called when reusable audio exists")

                main.create_black_video = fake_black_video
                main.mux_video_with_audio = fake_mux
                main.synthesize_audio_from_srt = fail_synthesize

                artifact = asyncio.run(main.create_video_artifact_for_file(db.get_file_by_id(file_record["id"]), "ko", voice_name="Kore", srt_source="original"))
        finally:
            db.DB_PATH = original_db_path
            db.JSON_DB_PATH = original_json_path
            main.mux_video_with_audio = original_mux
            main.create_black_video = original_black_video
            main.synthesize_audio_from_srt = original_synthesize

        self.assertTrue(artifact["metadata"]["reused_audio"])

    def test_captioned_dub_video_combines_dubbed_video_and_subtitles(self):
        original_db_path = db.DB_PATH
        original_json_path = db.JSON_DB_PATH
        original_create_video = main.create_video_artifact_for_file
        original_burn = main.burn_subtitles_into_video
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                db.DB_PATH = Path(temp_dir) / "files.sqlite"
                db.JSON_DB_PATH = Path(temp_dir) / "missing.json"
                db.init_db()
                dubbed_path = Path(temp_dir) / "dubbed.mp4"
                dubbed_path.write_bytes(b"dubbed")
                file_record = db.create_file_record(
                    "sample.srt",
                    "srt_project",
                    "text",
                    srt_text="1\n00:00:00,000 --> 00:00:01,000\nHello",
                )

                async def fake_create_video(file, language, voice_name=None, style_prompt=None, srt_source=None, tts_provider=None):
                    return db.create_artifact(
                        file["id"],
                        "video",
                        language,
                        str(dubbed_path),
                        "dubbed.mp4",
                        {"duration_ms": 1000, "voice_name": voice_name, "srt_source": srt_source or "original"},
                    )

                def fake_burn(source_video_path, srt_text, output_path, subtitle_style=None):
                    self.assertEqual(Path(source_video_path), dubbed_path)
                    self.assertIn("Hello", srt_text)
                    Path(output_path).write_bytes(b"captioned")
                    return {"font_size": 48, "position": "bottom"}

                main.create_video_artifact_for_file = fake_create_video
                main.burn_subtitles_into_video = fake_burn

                artifact = asyncio.run(
                    main.create_captioned_dub_video_artifact_for_file(
                        db.get_file_by_id(file_record["id"]),
                        "ko",
                        voice_name="Kore",
                        srt_source="original",
                    )
                )
        finally:
            db.DB_PATH = original_db_path
            db.JSON_DB_PATH = original_json_path
            main.create_video_artifact_for_file = original_create_video
            main.burn_subtitles_into_video = original_burn

        self.assertEqual(artifact["kind"], "captioned_dub_video")
        self.assertEqual(artifact["metadata"]["variant"], "captioned_dub")
        self.assertEqual(artifact["metadata"]["subtitle_style"]["position"], "bottom")

    def test_job_lifecycle_persists_status_and_metadata(self):
        original_db_path = db.DB_PATH
        original_json_path = db.JSON_DB_PATH
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                db.DB_PATH = Path(temp_dir) / "files.sqlite"
                db.JSON_DB_PATH = Path(temp_dir) / "missing.json"
                db.init_db()
                file_record = db.create_file_record("sample.srt", "srt_project", "text", srt_text="1\n00:00:00,000 --> 00:00:01,000\n안녕")
                job = db.create_job(file_record["id"], "audio", {"language": "ko", "voice_name": "Kore"})
                db.update_job(job["id"], status="running", progress=55, message="처리 중")
                db.update_job(job["id"], status="completed", progress=100, message="완료", result_artifact_id="artifact-1")
                detail = db.get_file_by_id(file_record["id"])
        finally:
            db.DB_PATH = original_db_path
            db.JSON_DB_PATH = original_json_path

        self.assertEqual(detail["job_summary"]["status"], "completed")
        self.assertEqual(detail["job_summary"]["progress"], 100)
        self.assertEqual(detail["jobs"][0]["metadata"]["voice_name"], "Kore")

    def test_correct_korean_srt_preserves_timing(self):
        class _FakeResponse:
            text = '[{"index": 1, "text": "튜토리얼입니다."}]'

        class _FakeModels:
            @staticmethod
            def generate_content(model, contents):
                return _FakeResponse()

        class _FakeGeminiClient:
            models = _FakeModels()

        original_client = main.gemini_text_client
        original_key = main.gemini_api_key
        main.gemini_text_client = _FakeGeminiClient()
        main.gemini_api_key = "test-key"
        try:
            output = asyncio.run(
                main.correct_korean_srt(
                    "1\n"
                    "00:00:00,000 --> 00:00:02,000\n"
                    "트토리얼입니다.\n"
                )
            )
        finally:
            main.gemini_text_client = original_client
            main.gemini_api_key = original_key

        self.assertIn("00:00:00,000 --> 00:00:02,000", output)
        self.assertIn("튜토리얼입니다.", output)


    def test_correct_korean_srt_accepts_resegmented_cues(self):
        class _FakeResponse:
            text = (
                '[{"start": "00:00:00,000", "end": "00:00:03,000", "text": "First natural cue."}, '
                '{"start": "00:00:03,000", "end": "00:00:06,000", "text": "Second natural cue."}]'
            )

        class _FakeModels:
            @staticmethod
            def generate_content(model, contents):
                return _FakeResponse()

        class _FakeGeminiClient:
            models = _FakeModels()

        original_client = main.gemini_text_client
        original_key = main.gemini_api_key
        main.gemini_text_client = _FakeGeminiClient()
        main.gemini_api_key = "test-key"
        try:
            output = asyncio.run(
                main.correct_korean_srt(
                    "1\n"
                    "00:00:00,000 --> 00:00:01,000\n"
                    "First\n\n"
                    "2\n"
                    "00:00:01,000 --> 00:00:06,000\n"
                    "natural cue. Second natural cue.\n"
                )
            )
        finally:
            main.gemini_text_client = original_client
            main.gemini_api_key = original_key

        self.assertIn("00:00:00,000 --> 00:00:03,000", output)
        self.assertIn("00:00:03,000 --> 00:00:06,000", output)
        self.assertIn("First natural cue.", output)
        self.assertIn("Second natural cue.", output)

    def _lecture_xlsx(self, rows):
        workbook = main.Workbook()
        sheet = workbook.active
        sheet.title = "slides"
        for row in rows:
            sheet.append(row)
        buffer = main.io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def test_parse_lecture_timeline_accepts_single_sheet_slide_scripts(self):
        content = self._lecture_xlsx([
            ["slide_no", "slide_file", "script"],
            [1, "LLM_slide_1.png", "Hello there"],
            [2, "LLM_slide_2.png", "friend"],
        ])
        result = main.parse_lecture_timeline_xlsx(
            content,
            ["LLM_slide_1.png", "LLM_slide_2.png"],
            "",
        )

        self.assertEqual(result["errors"], [])
        self.assertEqual(len(result["items"]), 2)
        self.assertEqual(result["items"][1]["slide_no"], 2)
        self.assertEqual(result["items"][1]["script"], "friend")
        self.assertEqual(result["items"][0]["script"], "Hello there")
        self.assertEqual(result["scripts"][0]["text"], "Hello there")

    def test_parse_lecture_timeline_reports_invalid_rows(self):
        content = self._lecture_xlsx([
            [1, "missing.png", "Hello"],
            [1, "LLM_slide_1.png", "Duplicate"],
            [2, "LLM_slide_2.png", ""],
            ["bad", "LLM_slide_2.png", "friend"],
        ])
        result = main.parse_lecture_timeline_xlsx(
            content,
            ["LLM_slide_1.png", "LLM_slide_2.png"],
            "",
        )

        self.assertTrue(any("missing.png" in error for error in result["errors"]))
        self.assertTrue(any("duplicate slide_no" in error for error in result["errors"]))
        self.assertTrue(any("slide_no must be a positive number" in error for error in result["errors"]))
        self.assertTrue(any("script is empty" in error for error in result["errors"]))

    def test_parse_lecture_timeline_warns_for_slide_number_gaps_and_unused_files(self):
        content = self._lecture_xlsx([
            ["slide_no", "slide_file", "script"],
            [1, "LLM_slide_1.png", "Hello"],
            [3, "LLM_slide_2.png", "friend"],
        ])
        result = main.parse_lecture_timeline_xlsx(
            content,
            ["LLM_slide_1.png", "LLM_slide_2.png", "unused.png"],
            "",
        )

        self.assertEqual(result["errors"], [])
        self.assertTrue(any("Gap before slide" in warning for warning in result["warnings"]))
        self.assertTrue(any("unused.png" in warning for warning in result["warnings"]))

    def test_lecture_template_uses_single_slide_script_sheet(self):
        workbook = load_workbook(main.io.BytesIO(main.create_lecture_timeline_template()), read_only=True, data_only=True)

        self.assertEqual(workbook.sheetnames, ["slides"])
        sheet = workbook["slides"]
        self.assertEqual([sheet["A1"].value, sheet["B1"].value, sheet["C1"].value], ["slide_no", "slide_file", "script"])
        self.assertTrue(str(sheet["C2"].value).strip())

    def test_align_generated_srt_uses_excel_script_with_actual_timings(self):
        generated_srt = (
            "1\n00:00:00,000 --> 00:00:02,000\n"
            "이번 챕터에서는 Gemina를 사용합니다.\n\n"
            "2\n00:00:02,000 --> 00:00:04,000\n"
            "이번 챕털의 로드맵입니다.\n"
        )
        timeline_items = [{
            "slide_no": 1,
            "script": "이번 챕터에서는 Gemini를 사용합니다. 이번 챕터의 로드맵입니다.",
            "speech_start_seconds": 0.0,
            "speech_end_seconds": 4.0,
        }]

        aligned_srt, metadata = main.align_generated_srt_to_lecture_scripts(generated_srt, timeline_items)

        self.assertTrue(metadata["aligned"])
        self.assertIn("Gemini", aligned_srt)
        self.assertIn("챕터", aligned_srt)
        self.assertNotIn("Gemina", aligned_srt)
        self.assertNotIn("챕털", aligned_srt)
        self.assertIn("00:00:00,000 --> 00:00:02,000", aligned_srt)
        self.assertIn("00:00:02,000 --> 00:00:04,000", aligned_srt)

    def test_align_generated_srt_does_not_pack_later_excel_sentences_into_early_cue(self):
        generated_srt = (
            "1\n00:00:00,000 --> 00:00:02,000\n"
            "이번 챕터에서는 Gemina와 구글 시트를 활용합니다.\n\n"
            "2\n00:00:02,000 --> 00:00:10,000\n"
            "여기서 말하는 결과물은 단순히 합계나 평균을 낸 표가 아니고요.\n\n"
            "3\n00:00:10,000 --> 00:00:18,000\n"
            "데이터를 볼 관점을 잡고 데이터 상태를 확인합니다.\n\n"
            "4\n00:00:18,000 --> 00:00:25,000\n"
            "핵심 지표 KPI를 고른 다음에 다음 행동 공유 자료로 정리하는 흐름입니다.\n"
        )
        timeline_items = [{
            "slide_no": 1,
            "script": (
                "이번 챕터에서는 Gemini와 Google Sheet를 활용합니다. "
                "여기서 말하는 결과물은 단순히 합계나 평균을 낸 표가 아니고요. "
                "데이터를 볼 관점을 잡고, 데이터 상태를 확인합니다. "
                "핵심 지표 KPI를 고른 다음에 다음 행동 공유 자료로 정리하는 흐름입니다."
            ),
            "speech_start_seconds": 0.0,
            "speech_end_seconds": 25.0,
        }]

        aligned_srt, _ = main.align_generated_srt_to_lecture_scripts(generated_srt, timeline_items)
        cues = main.parse_srt(aligned_srt)

        self.assertEqual(cues[1]["end"], 10.0)
        self.assertIn("여기서 말하는 결과물", cues[1]["text"])
        self.assertNotIn("데이터를 볼 관점", cues[1]["text"])
        self.assertGreaterEqual(cues[2]["start"], 10.0)
        self.assertIn("데이터를 볼 관점", cues[2]["text"])
        self.assertIn("KPI", cues[3]["text"])


if __name__ == "__main__":
    unittest.main()
