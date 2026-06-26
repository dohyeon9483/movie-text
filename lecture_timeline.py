import io
import re
from datetime import time as datetime_time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook


LECTURE_IMAGE_SLIDE_EXTENSIONS = {".png", ".jpg", ".jpeg"}
LECTURE_HTML_SLIDE_EXTENSIONS = {".html", ".htm"}
LECTURE_PDF_SLIDE_EXTENSIONS = {".pdf"}
LECTURE_SLIDE_EXTENSIONS = LECTURE_IMAGE_SLIDE_EXTENSIONS | LECTURE_HTML_SLIDE_EXTENSIONS | LECTURE_PDF_SLIDE_EXTENSIONS
LECTURE_SLIDE_HEADERS = {"slide_no", "slide_file", "script"}


def safe_upload_filename(filename: str) -> str:
    return Path(filename or "").name.replace("/", "_").replace("\\", "_").strip()


def parse_lecture_time_value(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime_time):
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
    if isinstance(value, timedelta):
        return value.total_seconds()
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return float(text)
    parts = text.split(":")
    if len(parts) not in {2, 3}:
        return None
    try:
        numbers = [float(part) for part in parts]
    except ValueError:
        return None
    if len(numbers) == 2:
        hours = 0
        minutes, seconds = numbers
    else:
        hours, minutes, seconds = numbers
    if minutes >= 60 or seconds >= 60:
        return None
    return hours * 3600 + minutes * 60 + seconds


def parse_positive_int(value: Any) -> Optional[int]:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def parse_lecture_timeline_xlsx(content: bytes, available_slide_files: List[str], srt_text: str = "") -> Dict:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Lecture XLSX could not be read: {exc}")

    sheet_names = {name.lower(): name for name in workbook.sheetnames}
    slides_sheet = workbook[sheet_names.get("slides", workbook.sheetnames[0])]
    available = {Path(name).name for name in available_slide_files}
    errors: List[str] = []
    warnings: List[str] = []
    items: List[Dict] = []
    scripts: List[Dict] = []
    seen_slide_numbers = set()
    used_slide_files = set()

    rows = list(slides_sheet.iter_rows(values_only=True))
    if not rows:
        errors.append("Slides sheet is empty.")

    first_values = [str(cell or "").strip().lower() for cell in (rows[0][:3] if rows else [])]
    has_header = set(first_values) >= LECTURE_SLIDE_HEADERS or any(value in LECTURE_SLIDE_HEADERS for value in first_values)
    data_rows = rows[1:] if has_header else rows

    for row_number, row in enumerate(data_rows, 2 if has_header else 1):
        values = list(row[:3])
        values.extend([None] * max(0, 3 - len(values)))
        if all(value is None or str(value).strip() == "" for value in values):
            continue

        raw_slide_no, raw_slide_file, raw_script = values
        slide_no = parse_positive_int(raw_slide_no)
        slide_file = safe_upload_filename(str(raw_slide_file or ""))
        script_text = str(raw_script or "").strip()

        if slide_no is None:
            errors.append(f"Row {row_number}: slide_no must be a positive number.")
            continue
        if slide_no in seen_slide_numbers:
            errors.append(f"Row {row_number}: duplicate slide_no {slide_no}.")
        seen_slide_numbers.add(slide_no)
        if not slide_file or slide_file not in available:
            errors.append(f"Row {row_number}: slide_file '{slide_file}' was not uploaded.")
        else:
            used_slide_files.add(slide_file)
        if not script_text:
            errors.append(f"Row {row_number}: script is empty.")

        items.append({
            "slide_no": slide_no,
            "slide_file": slide_file,
            "script": script_text,
        })
        if script_text:
            scripts.append({"slide_no": slide_no, "text": script_text})

    items.sort(key=lambda item: item["slide_no"])
    scripts.sort(key=lambda item: item["slide_no"])

    if not items:
        errors.append("Slides sheet has no valid slide rows.")
    if not scripts:
        errors.append("Slides sheet has no valid script text.")

    for previous, current in zip(items, items[1:]):
        if current["slide_no"] > previous["slide_no"] + 1:
            warnings.append(
                f"Gap before slide {current['slide_no']}: "
                f"slide {previous['slide_no'] + 1} to {current['slide_no'] - 1}."
            )

    used_base_files = {str(name).split("#", 1)[0] for name in used_slide_files}
    unused = sorted(
        name for name in available - used_slide_files
        if "#" not in name and name not in used_base_files
    )
    if unused:
        warnings.append(f"Uploaded slide files not used in XLSX: {', '.join(unused)}.")

    return {"items": items, "scripts": scripts, "errors": errors, "warnings": warnings, "has_header": has_header}


def create_lecture_timeline_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "slides"
    sheet.append(["slide_no", "slide_file", "script"])
    sheet.append([1, "LLM_slide_1.png", "안녕하세요. 오늘 강의의 핵심을 짧게 설명하겠습니다. 첫 번째 슬라이드에서는 전체 개요를 소개합니다."])
    sheet.append([2, "lecture_deck.html#1", "HTML 한 파일 안에 여러 장표가 있다면 파일명 뒤에 #1, #2처럼 장표 번호를 붙여 작성합니다."])
    sheet.append([3, "lecture_deck.html#2", "서버는 해당 HTML 장표를 1920x1080 PNG로 렌더링한 뒤 기존 영상 생성 흐름에 연결합니다."])
    sheet.append([4, "lecture_slides.pdf#1", "PDF 파일도 페이지 번호를 붙여 사용할 수 있습니다. 이미지, HTML, PDF 슬라이드를 섞어서 사용할 수 있습니다."])
    for column, width in {"A": 12, "B": 28, "C": 90}.items():
        sheet.column_dimensions[column].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
